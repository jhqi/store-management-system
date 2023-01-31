import ctypes
ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("myappid")
import resources_rc
from PyQt5.QtGui import QIcon
import sqlite3
import openpyxl
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QMainWindow, QMessageBox
from import_confirm import Ui_import_confirm
import var


class import_confirm_window(QMainWindow, Ui_import_confirm):
    def __init__(self):
        QMainWindow.__init__(self)
        self.setupUi(self)
        self.main_window = self
        self.setWindowIcon(QIcon(":/images/image.ico"))
        self.confirm_btn.clicked.disconnect()
        self.confirm_btn.clicked.connect(self.on_confirm_btn_clicked)
        self.cancel_btn.clicked.connect(self.on_cancel_btn_clicked)
        if var.import_flag == 1:
            self.label.setText("警告！导入前将删除全部现有存货数据！")
            self.label_2.setText("该操作将删除现有存货数据！")
        elif var.import_flag == 2:
            self.label.setText("警告！导入前将删除全部现有入库记录！")
            self.label_2.setText("该操作将删除现有入库记录！")
        else:
            self.label.setText("警告！导入前将删除全部现有出库记录！")
            self.label_2.setText("该操作将删除现有出库记录！")
        self.label_2.setStyleSheet('background-color:red')

    def on_cancel_btn_clicked(self):
        self.close()

    def on_confirm_btn_clicked(self):
        try:
            if var.import_flag == 1:
                t_path = QtWidgets.QFileDialog.getOpenFileName(self, "请选择要读取的文件", "C:/", "Excel表格 (*.xlsx)")
                if t_path[0] == "":
                    return
                else:
                    file_path = t_path[0]
                conn = sqlite3.connect("material_management.db")
                conn.text_factory = str
                cur = conn.cursor()
                sql = "drop table material"
                cur.execute(sql)
                conn.commit()
                cur.execute(
                    "create table material(material_id varchar(32) primary key,material_name varchar(100),spec varchar(100),unit varchar(10), per_price float, pos varchar(100), ori_num float,now_num float)")
                conn.commit()
                cur.close()
                conn.close()

                workbook = openpyxl.load_workbook(file_path)
                table = workbook.active
                rows = table.max_row
                cols = table.max_column
                all_ll = []
                for row in range(1, rows):
                    tmp = []
                    for col in range(0, cols):
                        data = str(table.cell(row=row + 1, column=col + 1).value)
                        if col == 4:
                            if data == '未知':
                                data = '-1'
                        tmp.append(data)
                    all_ll.append(tmp)
                conn = sqlite3.connect("material_management.db")
                conn.text_factory = str
                cur = conn.cursor()
                sql = "insert into material values "
                for i in range(0, len(all_ll)):
                    sql += "('" + all_ll[i][0] + "','" + all_ll[i][1] + "','" + all_ll[i][2] + "','" + all_ll[i][
                        3] + "'," + all_ll[i][4] + ",'" + all_ll[i][5] + "'," + all_ll[i][6] + "," + all_ll[i][7] + ")"
                    if i < len(all_ll) - 1:
                        sql += ","
                cur.execute(sql)
                conn.commit()
                cur.close()
                conn.close()
                QMessageBox.question(self, '导入存货成功！', '存货数据已全部导入数据库！', QMessageBox.Yes)

            elif var.import_flag == 2:
                t_path = QtWidgets.QFileDialog.getOpenFileName(self, "请选择要读取的文件", "C:/", "Excel表格 (*.xlsx)")
                if t_path[0] == "":
                    return
                else:
                    file_path = t_path[0]
                conn = sqlite3.connect("material_management.db")
                conn.text_factory = str
                cur = conn.cursor()
                sql = "drop table in_log"
                cur.execute(sql)
                conn.commit()
                cur.execute(
                    "create table in_log(date_time int,material_id varchar(32),material_name varchar(100),spec varchar(100),in_num float, per_price float, pos varchar(100), total_price float,user_man varchar(20),agree_man varchar(20), in_log_id INTEGER PRIMARY KEY AUTOINCREMENT)")

                conn.commit()
                cur.close()
                conn.close()

                workbook = openpyxl.load_workbook(file_path)
                table = workbook.active
                rows = table.max_row
                cols = table.max_column
                all_ll = []
                for row in range(1, rows):
                    tmp = []
                    for col in range(0, cols):
                        data = str(table.cell(row=row + 1, column=col + 1).value)
                        if col == 0:
                            data = data[0:10]
                            ll = data.split('-')
                            data = "".join(ll)
                        if col == 5:
                            if data == '未知':
                                data = '-1'
                        if col == 7:
                            if data == '未知':
                                data = '-1'
                        tmp.append(data)
                    all_ll.append(tmp)
                conn = sqlite3.connect("material_management.db")
                conn.text_factory = str
                cur = conn.cursor()
                sql = "insert into in_log values "
                for i in range(0, len(all_ll)):
                    sql += "(" + all_ll[i][0] + ",'" + all_ll[i][1] + "','" + all_ll[i][2] + "','" + all_ll[i][
                        3] + "'," + all_ll[i][4] + "," + all_ll[i][5] + ",'" + all_ll[i][6] + "'," + all_ll[i][
                               7] + ",'" + all_ll[i][8] + "','" + all_ll[i][9] + "',null)"
                    if i < len(all_ll) - 1:
                        sql += ","
                cur.execute(sql)
                conn.commit()
                cur.close()
                conn.close()
                QMessageBox.question(self, '导入入库记录成功！', '入库记录已全部导入数据库！', QMessageBox.Yes)

            else:
                t_path = QtWidgets.QFileDialog.getOpenFileName(self, "请选择要读取的文件", "C:/", "Excel表格 (*.xlsx)")
                if t_path[0] == "":
                    return
                else:
                    file_path = t_path[0]
                conn = sqlite3.connect("material_management.db")
                conn.text_factory = str
                cur = conn.cursor()
                sql = "drop table out_log"
                cur.execute(sql)
                conn.commit()
                cur.execute(
                    "create table out_log(date_time int,material_id varchar(32),material_name varchar(100),spec varchar(100),out_num float,per_price float, pos varchar(100), total_price float,user_man varchar(20),agree_man varchar(20), out_log_id INTEGER PRIMARY KEY AUTOINCREMENT)")

                conn.commit()
                cur.close()
                conn.close()

                workbook = openpyxl.load_workbook(file_path)
                table = workbook.active
                rows = table.max_row
                cols = table.max_column
                all_ll = []
                for row in range(1, rows):
                    tmp = []
                    for col in range(0, cols):
                        data = str(table.cell(row=row + 1, column=col + 1).value)
                        if col == 0:
                            data = data[0:10]
                            ll = data.split('-')
                            data = "".join(ll)
                        if col == 5:
                            if data == '未知':
                                data = '-1'
                        if col == 7:
                            if data == '未知':
                                data = '-1'
                        tmp.append(data)
                    all_ll.append(tmp)
                conn = sqlite3.connect("material_management.db")
                conn.text_factory = str
                cur = conn.cursor()
                sql = "insert into out_log values "
                for i in range(0, len(all_ll)):
                    sql += "(" + all_ll[i][0] + ",'" + all_ll[i][1] + "','" + all_ll[i][2] + "','" + all_ll[i][
                        3] + "'," + \
                           all_ll[i][4] + "," + all_ll[i][5] + ",'" + all_ll[i][6] + "'," + all_ll[i][7] + ",'" + \
                           all_ll[i][
                               8] + "','" + all_ll[i][9] + "',null)"
                    if i < len(all_ll) - 1:
                        sql += ","
                cur.execute(sql)
                conn.commit()
                cur.close()
                conn.close()
                QMessageBox.question(self, '导入出库记录成功！', '出库记录已全部导入数据库！', QMessageBox.Yes)
        except:
            QMessageBox.question(self, '导入失败！', 'Excel数据格式不符合规范！', QMessageBox.Yes)
            self.close()
        self.close()
